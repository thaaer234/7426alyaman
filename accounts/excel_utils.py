"""
Excel Export Utilities for Financial Reports
Provides comprehensive Excel export functionality with proper formatting
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from decimal import Decimal
from datetime import datetime, date
from django.http import HttpResponse
from django.utils import timezone
import locale


class ExcelFormatter:
    """Handles Excel formatting for financial reports"""
    
    # Color scheme for reports
    HEADER_COLOR = "366092"  # Dark blue
    SUBHEADER_COLOR = "5B9BD5"  # Light blue
    TOTAL_COLOR = "70AD47"  # Green
    ALTERNATE_COLOR = "F2F2F2"  # Light gray
    
    def __init__(self, workbook):
        self.workbook = workbook
        self.setup_styles()
    
    def setup_styles(self):
        """Setup common styles for the workbook"""
        # Header style
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Subheader style
        self.subheader_font = Font(name='Arial', size=11, bold=True, color='000000')
        self.subheader_fill = PatternFill(start_color=self.SUBHEADER_COLOR, end_color=self.SUBHEADER_COLOR, fill_type='solid')
        self.subheader_alignment = Alignment(horizontal='center', vertical='center')
        
        # Total style
        self.total_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.total_fill = PatternFill(start_color=self.TOTAL_COLOR, end_color=self.TOTAL_COLOR, fill_type='solid')
        self.total_alignment = Alignment(horizontal='right', vertical='center')
        
        # Data style
        self.data_font = Font(name='Arial', size=10)
        self.data_alignment = Alignment(horizontal='right', vertical='center')
        
        # Border style
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def format_header(self, worksheet, row, start_col, end_col, text):
        """Format header row"""
        cell = worksheet.cell(row=row, column=start_col, value=text)
        worksheet.merge_cells(f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}")
        
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment
        cell.border = self.thin_border
        
        return cell
    
    def format_subheader(self, worksheet, row, start_col, end_col, text):
        """Format subheader row"""
        cell = worksheet.cell(row=row, column=start_col, value=text)
        worksheet.merge_cells(f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}")
        
        cell.font = self.subheader_font
        cell.fill = self.subheader_fill
        cell.alignment = self.subheader_alignment
        cell.border = self.thin_border
        
        return cell
    
    def format_total_row(self, worksheet, row, start_col, end_col, text):
        """Format total row"""
        cell = worksheet.cell(row=row, column=start_col, value=text)
        worksheet.merge_cells(f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}")
        
        cell.font = self.total_font
        cell.fill = self.total_fill
        cell.alignment = self.total_alignment
        cell.border = self.thin_border
        
        return cell
    
    def format_data_cell(self, worksheet, row, col, value, is_number=False):
        """Format data cell"""
        cell = worksheet.cell(row=row, column=col, value=value)
        cell.font = self.data_font
        cell.border = self.thin_border
        
        if is_number:
            cell.alignment = self.data_alignment
            cell.number_format = '#,##0.00'
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        return cell
    
    def format_currency_cell(self, worksheet, row, col, value):
        """Format currency cell with comma separators"""
        cell = worksheet.cell(row=row, column=col, value=float(value) if value else 0)
        cell.font = self.data_font
        cell.alignment = self.data_alignment
        cell.border = self.thin_border
        cell.number_format = '#,##0.00'
        
        return cell
    
    def auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width


class FinancialReportExporter:
    """Main class for exporting financial reports to Excel"""
    
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.formatter = ExcelFormatter(self.workbook)
    
    def create_cost_center_analysis_report(self, cost_centers_data, period_start=None, period_end=None):
        """Create Cost Center Analysis Report"""
        worksheet = self.workbook.active
        worksheet.title = "Cost Center Analysis"
        
        # Report header
        current_row = 1
        self.formatter.format_header(worksheet, current_row, 1, 8, 
                                   "تقرير تحليل مراكز التكلفة - Cost Center Analysis Report")
        current_row += 2
        
        # Period information
        if period_start and period_end:
            period_text = f"الفترة من {period_start} إلى {period_end} - Period: {period_start} to {period_end}"
            self.formatter.format_subheader(worksheet, current_row, 1, 8, period_text)
            current_row += 2
        
        # Column headers
        headers = [
            "رمز المركز", "اسم المركز", "إجمالي المصروفات", "رواتب المدرسين",
            "مصروفات أخرى", "عدد الدورات", "إجمالي الإيرادات", "الربح/الخسارة"
        ]
        
        for col, header in enumerate(headers, 1):
            self.formatter.format_subheader(worksheet, current_row, col, col, header)
        current_row += 1
        
        # Data rows
        total_expenses = Decimal('0')
        total_teacher_salaries = Decimal('0')
        total_other_expenses = Decimal('0')
        total_revenue = Decimal('0')
        total_profit_loss = Decimal('0')
        
        for cost_center_data in cost_centers_data:
            # Cost center code and name
            self.formatter.format_data_cell(worksheet, current_row, 1, cost_center_data['code'])
            self.formatter.format_data_cell(worksheet, current_row, 2, cost_center_data['name'])
            
            # Financial data
            expenses = cost_center_data.get('total_expenses', Decimal('0'))
            teacher_salaries = cost_center_data.get('teacher_salaries', Decimal('0'))
            other_expenses = cost_center_data.get('other_expenses', Decimal('0'))
            revenue = cost_center_data.get('total_revenue', Decimal('0'))
            profit_loss = revenue - expenses
            
            self.formatter.format_currency_cell(worksheet, current_row, 3, expenses)
            self.formatter.format_currency_cell(worksheet, current_row, 4, teacher_salaries)
            self.formatter.format_currency_cell(worksheet, current_row, 5, other_expenses)
            self.formatter.format_currency_cell(worksheet, current_row, 6, cost_center_data.get('course_count', 0))
            self.formatter.format_currency_cell(worksheet, current_row, 7, revenue)
            self.formatter.format_currency_cell(worksheet, current_row, 8, profit_loss)
            
            # Update totals
            total_expenses += expenses
            total_teacher_salaries += teacher_salaries
            total_other_expenses += other_expenses
            total_revenue += revenue
            total_profit_loss += profit_loss
            
            current_row += 1
        
        # Total row
        current_row += 1
        self.formatter.format_total_row(worksheet, current_row, 1, 2, "المجموع الكلي - Total")
        self.formatter.format_currency_cell(worksheet, current_row, 3, total_expenses)
        self.formatter.format_currency_cell(worksheet, current_row, 4, total_teacher_salaries)
        self.formatter.format_currency_cell(worksheet, current_row, 5, total_other_expenses)
        self.formatter.format_data_cell(worksheet, current_row, 6, sum(cc.get('course_count', 0) for cc in cost_centers_data))
        self.formatter.format_currency_cell(worksheet, current_row, 7, total_revenue)
        self.formatter.format_currency_cell(worksheet, current_row, 8, total_profit_loss)
        
        # Auto-adjust columns
        self.formatter.auto_adjust_columns(worksheet)
        
        return self.workbook
    
    def create_cost_center_cash_flow_report(self, cash_flow_data, period_start=None, period_end=None):
        """Create Cost Center Cash Flow Report"""
        worksheet = self.workbook.active
        worksheet.title = "Cost Center Cash Flow"
        
        # Report header
        current_row = 1
        self.formatter.format_header(worksheet, current_row, 1, 7, 
                                   "تقرير التدفق النقدي لمراكز التكلفة - Cost Center Cash Flow Report")
        current_row += 2
        
        # Period information
        if period_start and period_end:
            period_text = f"الفترة من {period_start} إلى {period_end} - Period: {period_start} to {period_end}"
            self.formatter.format_subheader(worksheet, current_row, 1, 7, period_text)
            current_row += 2
        
        # Column headers
        headers = [
            "رمز المركز", "اسم المركز", "التدفق الداخل", "التدفق الخارج",
            "الرصيد الافتتاحي", "الرصيد الختامي", "الوضع المالي"
        ]
        
        for col, header in enumerate(headers, 1):
            self.formatter.format_subheader(worksheet, current_row, col, col, header)
        current_row += 1
        
        # Data rows
        total_inflow = Decimal('0')
        total_outflow = Decimal('0')
        total_opening_balance = Decimal('0')
        total_closing_balance = Decimal('0')
        
        for cash_flow_item in cash_flow_data:
            # Cost center code and name
            self.formatter.format_data_cell(worksheet, current_row, 1, cash_flow_item['code'])
            self.formatter.format_data_cell(worksheet, current_row, 2, cash_flow_item['name'])
            
            # Cash flow data
            inflow = cash_flow_item.get('inflow', Decimal('0'))
            outflow = cash_flow_item.get('outflow', Decimal('0'))
            opening_balance = cash_flow_item.get('opening_balance', Decimal('0'))
            closing_balance = opening_balance + inflow - outflow
            
            # Determine financial position
            if closing_balance > 0:
                position = "رصيد موجب - Positive"
            elif closing_balance < 0:
                position = "رصيد سالب - Negative"
            else:
                position = "متوازن - Balanced"
            
            self.formatter.format_currency_cell(worksheet, current_row, 3, inflow)
            self.formatter.format_currency_cell(worksheet, current_row, 4, outflow)
            self.formatter.format_currency_cell(worksheet, current_row, 5, opening_balance)
            self.formatter.format_currency_cell(worksheet, current_row, 6, closing_balance)
            self.formatter.format_data_cell(worksheet, current_row, 7, position)
            
            # Update totals
            total_inflow += inflow
            total_outflow += outflow
            total_opening_balance += opening_balance
            total_closing_balance += closing_balance
            
            current_row += 1
        
        # Total row
        current_row += 1
        self.formatter.format_total_row(worksheet, current_row, 1, 2, "المجموع الكلي - Total")
        self.formatter.format_currency_cell(worksheet, current_row, 3, total_inflow)
        self.formatter.format_currency_cell(worksheet, current_row, 4, total_outflow)
        self.formatter.format_currency_cell(worksheet, current_row, 5, total_opening_balance)
        self.formatter.format_currency_cell(worksheet, current_row, 6, total_closing_balance)
        
        # Overall position
        if total_closing_balance > 0:
            overall_position = "رصيد موجب - Positive"
        elif total_closing_balance < 0:
            overall_position = "رصيد سالب - Negative"
        else:
            overall_position = "متوازن - Balanced"
        self.formatter.format_data_cell(worksheet, current_row, 7, overall_position)
        
        # Auto-adjust columns
        self.formatter.auto_adjust_columns(worksheet)
        
        return self.workbook
    
    def create_comprehensive_financial_report(self, analysis_data, cash_flow_data, period_start=None, period_end=None):
        """Create comprehensive financial report with multiple sheets"""
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Create Cost Center Analysis sheet
        analysis_sheet = self.workbook.create_sheet("Cost Center Analysis")
        self.formatter = ExcelFormatter(self.workbook)
        
        # Create Cash Flow sheet
        cash_flow_sheet = self.workbook.create_sheet("Cash Flow Analysis")
        
        # Generate both reports
        self.create_cost_center_analysis_report(analysis_data, period_start, period_end)
        self.create_cost_center_cash_flow_report(cash_flow_data, period_start, period_end)
        
        return self.workbook


def format_number_with_commas(value):
    """Format number with comma separators"""
    if value is None:
        return "0.00"
    
    try:
        # Convert to float first, then format
        float_value = float(value)
        return f"{float_value:,.2f}"
    except (ValueError, TypeError):
        return "0.00"


def create_excel_response(workbook, filename):
    """Create HTTP response for Excel file download"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    workbook.save(response)
    return response


def get_period_display(period_start, period_end):
    """Get formatted period display text"""
    if period_start and period_end:
        return f"{period_start} - {period_end}"
    elif period_start:
        return f"من {period_start}"
    elif period_end:
        return f"حتى {period_end}"
    else:
        return "جميع الفترات - All Periods"
